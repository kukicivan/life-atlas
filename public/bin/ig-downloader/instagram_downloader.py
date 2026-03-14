#!/usr/bin/env python3
"""
Instagram Photo Downloader
Auth: cookies file (Netscape format, exported from Chrome)
Backend: gallery-dl

Output layout:
  instagram_downloads/
  └── <profile>/
      ├── photos/                  ← downloaded images
      ├── metadata/                ← one .json per image (gallery-dl metadata)
      └── <profile>_comments.xlsx  ← comments spreadsheet (when download_comments=True)
"""

import os
import sys
import json
import shutil
import subprocess
import tempfile
from pathlib import Path

# ============================================================================
# CONFIG  (loaded from config.json next to this script)
# ============================================================================

def load_config():
    config_path = Path(__file__).parent / "config.json"
    if not config_path.exists():
        print("ERROR: config.json not found.")
        print("  Copy config.example.json to config.json and fill in your values.")
        sys.exit(1)
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)

CONFIG = load_config()

# ============================================================================
# CHECKS
# ============================================================================

def check_gallery_dl():
    try:
        r = subprocess.run(["gallery-dl", "--version"], capture_output=True, text=True)
        print(f"gallery-dl {r.stdout.strip()}")
    except FileNotFoundError:
        print("ERROR: gallery-dl not found.")
        print("  Install: pip install gallery-dl")
        sys.exit(1)


def check_cookies(path):
    if not Path(path).exists():
        print(f"ERROR: Cookie file not found: {path}")
        print("  Export from Chrome with the 'Get cookies.txt LOCALLY' extension,")
        print("  then save it next to this script as www.instagram.com_cookies.txt")
        sys.exit(1)
    print(f"Cookies:  {path}")


# ============================================================================
# DOWNLOAD
# ============================================================================

def run_download(cfg):
    """
    Run gallery-dl into the photos/ subfolder, then move all .json metadata
    files into the metadata/ subfolder.
    Returns True on success.
    """
    target = cfg["target_profile"]
    profile_folder = Path(cfg["download_folder"]) / target
    photos_folder   = profile_folder / "photos"
    metadata_folder = profile_folder / "metadata"

    photos_folder.mkdir(parents=True, exist_ok=True)
    metadata_folder.mkdir(parents=True, exist_ok=True)

    # Minimal gallery-dl config (added on top of the user's own config)
    gd_config = {}
    if cfg["download_comments"]:
        gd_config = {
            "extractor": {
                "instagram": {
                    "comments": True   # fetch actual comment text from the API
                }
            }
        }

    config_file = None
    try:
        if gd_config:
            tmp = tempfile.NamedTemporaryFile(
                mode="w", suffix=".json", delete=False, encoding="utf-8"
            )
            json.dump(gd_config, tmp)
            tmp.close()
            config_file = tmp.name

        cmd = ["gallery-dl"]

        if config_file:
            cmd.extend(["--config", config_file])

        cmd.extend([
            "--cookies", cfg["cookies_file"],
            "--directory", str(photos_folder),        # exact output dir, no sub-dirs
            "--filename", "{shortcode}_{num}.{extension}",
            "--write-metadata",                        # writes {filename}.json alongside each image
        ])

        if cfg["max_photos"] > 0:
            cmd.extend(["--range", f'1-{cfg["max_photos"]}'])

        if cfg["skip_videos"]:
            cmd.extend(["--filter", 'extension not in ("mp4", "mov", "webm")'])

        cmd.append(f"https://www.instagram.com/{target}/")

        print(f"\nRunning gallery-dl...\n")
        result = subprocess.run(cmd)

        if result.returncode != 0:
            return False

        # Move .json files from photos/ to metadata/
        for json_file in photos_folder.glob("*.json"):
            dest = metadata_folder / json_file.name
            shutil.move(str(json_file), str(dest))

        return True

    finally:
        if config_file and os.path.exists(config_file):
            os.unlink(config_file)


# ============================================================================
# EXCEL EXPORT
# ============================================================================

def build_excel(cfg):
    """
    Parse .json files from metadata/ and write one Excel row per comment.
    Posts with no comments get one row with empty comment columns.

    Output: instagram_downloads/<profile>/<profile>_comments.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return

    target = cfg["target_profile"]
    profile_folder  = Path(cfg["download_folder"]) / target
    metadata_folder = profile_folder / "metadata"

    json_files = sorted(metadata_folder.glob("*.json"))
    if not json_files:
        print("No metadata JSON files found - skipping Excel export.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comments"

    headers = [
        "image_filename",
        "shortcode",
        "post_url",
        "caption",
        "likes",
        "post_date",
        "comment_author",
        "comment_text",
        "comment_date",
    ]
    ws.append(headers)

    # Style the header row
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Freeze header row; keep image_filename always visible
    ws.freeze_panes = "B2"

    row_count = 0
    for json_file in json_files:
        try:
            meta = json.loads(json_file.read_text(encoding="utf-8"))
        except Exception:
            continue

        # json_file.name is e.g. "ABC123_1.jpg.json"
        # json_file.stem strips the last ".json" -> "ABC123_1.jpg"
        image_filename = json_file.stem
        shortcode = meta.get("shortcode", "")
        post_url  = meta.get("post_url", f"https://instagram.com/p/{shortcode}/")
        caption   = meta.get("description", "")
        likes     = meta.get("likes", "")
        post_date = str(meta.get("date", ""))

        comments = meta.get("comments", [])

        if isinstance(comments, list) and comments:
            for c in comments:
                author = c.get("username", "")
                text   = c.get("text", "")
                cdate  = str(c.get("date", ""))
                ws.append([
                    image_filename, shortcode, post_url, caption,
                    likes, post_date, author, text, cdate,
                ])
                row_count += 1
        else:
            # No comments - one row so the image still appears in the sheet
            ws.append([
                image_filename, shortcode, post_url, caption,
                likes, post_date, "", "", "",
            ])
            row_count += 1

    # Approximate column widths
    column_widths = [30, 14, 50, 60, 8, 22, 20, 80, 22]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    excel_path = profile_folder / f"{target}_comments.xlsx"
    wb.save(excel_path)
    print(f"Excel saved:  {excel_path}")
    print(f"Rows written: {row_count}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("\n" + "=" * 58)
    print("  INSTAGRAM PHOTO DOWNLOADER".center(58))
    print("  gallery-dl + cookies auth".center(58))
    print("=" * 58 + "\n")

    if not CONFIG["target_profile"]:
        print("ERROR: Set target_profile in CONFIG")
        sys.exit(1)

    check_gallery_dl()
    check_cookies(CONFIG["cookies_file"])

    target = CONFIG["target_profile"]
    print(f"Target:   @{target}")
    print(f"Max:      {CONFIG['max_photos'] or 'all'}")
    print(f"Comments: {'yes (will save to Excel)' if CONFIG['download_comments'] else 'no'}")
    print(f"Output:   instagram_downloads/{target}/")
    print(f"          |- photos/")
    print(f"          |- metadata/")
    if CONFIG["download_comments"]:
        print(f"          `- {target}_comments.xlsx")

    success = run_download(CONFIG)

    if not success:
        print("\nDownload failed (gallery-dl returned a non-zero exit code).")
        sys.exit(1)

    if CONFIG["download_comments"]:
        print("\nBuilding Excel file with comments...")
        build_excel(CONFIG)

    print("\nDone!")


if __name__ == "__main__":
    main()
